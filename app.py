import streamlit as st
import pandas as pd
import numpy as np
import numpy_financial as npf
import re
import plotly.express as px
import os
import math
from io import BytesIO
# --------------------------------------------------------------------------
# [설정] 페이지 기본
# --------------------------------------------------------------------------
st.set_page_config(page_title="공식 배관 투자 결재 시스템 (Pipeline Approval)", page_icon="🏗️", layout="wide")
# --------------------------------------------------------------------------
# [함수] 금융 계산 로직
# --------------------------------------------------------------------------
def manual_npv(rate, values):
    return sum(v / ((1 + rate) ** i) for i, v in enumerate(values))
def excel_round(val):
    return math.floor(val + 0.5) if val > 0 else math.ceil(val - 0.5)
def calculate_simulation(sim_len, sim_inv, sim_contrib, sim_other, sim_vol, sim_rev, sim_cost,
                         sim_jeon, sim_basic_rev, rate, tax, dep_period, analysis_period, c_maint, c_adm_jeon, c_adm_m):

    net_inv = excel_round(sim_inv - sim_contrib - sim_other)
    margin_total = excel_round((sim_rev - sim_cost) + sim_basic_rev)

    cost_sga = excel_round((sim_len * c_maint) + (sim_len * c_adm_m) + (sim_jeon * c_adm_jeon))
    annual_depreciation = excel_round(sim_inv / dep_period) if dep_period > 0 else 0

    ebit = margin_total - cost_sga - annual_depreciation
    net_income = excel_round(ebit * (1 - tax))
    fixed_ocf = net_income + annual_depreciation

    flows = [-net_inv]
    ocfs = []

    for year in range(1, int(analysis_period) + 1):
        flows.append(fixed_ocf)
        ocfs.append(fixed_ocf)
    npv_val = manual_npv(rate, flows)
    irr_val = None
    irr_reason = ""

    if net_inv <= 0:
        irr_reason = "초기 투자비 ≤ 0"
    elif all(f <= 0 for f in ocfs):
        irr_reason = "운영 적자 지속"
    else:
        try:
            irr_val = npf.irr(flows)
        except:
            irr_reason = "계산 오류"

    return npv_val, irr_val, irr_reason, flows
def get_col_idx(df, keywords, exact=False):
    for col_idx in range(df.shape[1]):
        for row_idx in range(min(20, df.shape[0])):
            val = str(df.iloc[row_idx, col_idx]).replace(" ", "").replace("\n", "")
            for kw in keywords:
                if exact:
                    if val == kw: return col_idx
                else:
                    if kw in val: return col_idx
    return None
# --------------------------------------------------------------------------
# [UI] 공통 사이드바 (파일 업로드 및 자동 스캔)
# --------------------------------------------------------------------------
with st.sidebar:
    st.title("메뉴 네비게이션")
    menu_choice = st.radio(
        "이동할 페이지를 선택하세요:",
        ('1. 배관 투자 경제성 결재 대시보드', '2. 배관 투자 승인 내역', '3. 품의서 결재')
    )
    st.divider()
    st.header("📂 데이터 로드 (자동/수동)")
    st.markdown("깃허브에 올린 엑셀 파일을 자동으로 찾아냅니다. (급할 땐 아래에 바로 드래그 업로드 가능)")

    uploaded_files = st.file_uploader("수동 파일 업로드 (기본값 덮어쓰기용)", accept_multiple_files=True, type=['csv', 'xlsx', 'xls'])

    working_files = []

    if uploaded_files:
        working_files = uploaded_files
        st.success(f"📌 수동 업로드 모드: {len(working_files)}개 파일 적용됨")
    else:
        scanned_files = []
        for root, dirs, files in os.walk("."):
            if '.git' in root or '.streamlit' in root:
                continue

            for f in files:
                if f.endswith(('.csv', '.xlsx', '.xls')) and not f.startswith('~'):
                    if '승인내역' in f:
                        continue
                    if '기초자료' in f or '현황' in f or '투자' in f:
                        scanned_files.append(os.path.join(root, f))

        working_files = list(set(scanned_files))

        if working_files:
            st.info(f"💡 자동 로드 모드: 깃허브에서 {len(working_files)}개 엑셀 파일 발견!")
            for wf in working_files:
                st.caption(f"✅ {os.path.basename(wf)}")
        else:
            st.warning("⚠️ 깃허브에서 엑셀 파일을 찾지 못했습니다. 깃허브에 파일을 올리거나 수동 업로드 하세요.")

    st.divider()
# ==========================================================================
# 탭 1: 기존 배관 투자 경제성 결재 대시보드
# ==========================================================================
if menu_choice == '1. 배관 투자 경제성 결재 대시보드':
    with st.sidebar:
        st.header("⚙️ 분석 변수 설정")
        st.info("💡 전산 시스템의 NPV와 일치하도록 아래 단가와 세율을 시스템 세팅값과 동일하게 맞춰주세요.")

        rate_pct = st.number_input("할인율 (%)", value=5.86, step=0.01, format="%.2f")
        tax_pct = st.number_input("법인세율+주민세율 (%)", value=22.0, step=0.1, format="%.1f")
        dep_period = st.number_input("감가상각 연수 (년)", value=30, step=1)
        analysis_period = st.number_input("경제성 분석 연수 (년)", value=30, step=1)

        st.subheader("💰 비용 단가")
        c_maint = st.number_input("유지비 (원/m)", value=8337, format="%d", key="input_c_maint")
        c_adm_jeon = st.number_input("관리비 (원/전)", value=17357, format="%d", key="input_c_adm_jeon")
        c_adm_m = st.number_input("관리비 (원/m)", value=11870, format="%d", key="input_c_adm_m")
        sim_basic_price = st.number_input("주택용 월 기본요금 단가 (원)", value=900, step=10, format="%d", key="input_sim_basic_price")
        RATE = rate_pct / 100
        TAX = tax_pct / 100
    st.title("🏗️ 배관 투자 경제성 결재 대시보드")
    st.markdown("전산 시스템 Raw 데이터를 업로드하여 경제성을 시뮬레이션합니다. **분석에서 제외할 항목은 체크 해제**하세요.")
    if working_files:
        clean_df_list = []
        for file_obj in working_files:
            try:
                file_name = file_obj.name if hasattr(file_obj, 'name') else os.path.basename(file_obj)

                # 현황정리 파일 또는 승인내역/품의서 파일은 탭1에서 제외
                if '현황' in file_name or '승인내역' in file_name or '품의서' in file_name:
                    continue
                if '투자' in file_name and '기초자료' not in file_name:
                    continue

                if hasattr(file_obj, 'seek'):
                    file_obj.seek(0)

                match = re.search(r'(\d+)차', file_name)
                cha_num = int(match.group(1)) if match else 1
                if file_name.endswith('.csv'):
                    df = pd.read_csv(file_obj, header=None, encoding='utf-8-sig')
                else:
                    df = pd.read_excel(file_obj, header=None)
                df = df.iloc[:100]
                idx_usage = 0
                idx_name = get_col_idx(df, ["구간명"], exact=True)
                idx_len = get_col_idx(df, ["길이(m)", "배관길이"], exact=False)
                idx_inv = get_col_idx(df, ["배관투자금액", "총공사비"], exact=False)
                idx_contrib = get_col_idx(df, ["시설분담금"], exact=False)
                idx_other = get_col_idx(df, ["기타이익", "보조금"], exact=False)
                idx_jeon = get_col_idx(df, ["수요전수계", "총전수"], exact=False)
                idx_jeon_apt = get_col_idx(df, ["공동주택전수"], exact=False)
                idx_jeon_single = get_col_idx(df, ["단독주택전수"], exact=False)
                idx_vol = get_col_idx(df, ["계(MJ)"], exact=False)
                idx_rev = get_col_idx(df, ["연간판매액", "판매액"], exact=False)
                idx_cost = get_col_idx(df, ["연간판매원가", "판매원가"], exact=False)

                idx_npv = get_col_idx(df, ["NPV"], exact=False)
                idx_irr = get_col_idx(df, ["IRR"], exact=False)
                if idx_name is None:
                    continue
                mapped_data = {}
                mapped_data['차수'] = cha_num
                mapped_data['용도'] = df.iloc[:, idx_usage] if idx_usage is not None else '미분류'
                mapped_data['구간명'] = df.iloc[:, idx_name]
                mapped_data['길이'] = df.iloc[:, idx_len] if idx_len is not None else 0
                mapped_data['투자비'] = df.iloc[:, idx_inv] if idx_inv is not None else 0
                mapped_data['분담금'] = df.iloc[:, idx_contrib] if idx_contrib is not None else 0
                mapped_data['기타이익'] = df.iloc[:, idx_other] if idx_other is not None else 0
                mapped_data['총전수'] = df.iloc[:, idx_jeon] if idx_jeon is not None else 0
                mapped_data['공동주택전수'] = df.iloc[:, idx_jeon_apt] if idx_jeon_apt is not None else 0
                mapped_data['단독주택전수'] = df.iloc[:, idx_jeon_single] if idx_jeon_single is not None else 0
                mapped_data['판매량'] = df.iloc[:, idx_vol] if idx_vol is not None else 0
                mapped_data['판매액'] = df.iloc[:, idx_rev] if idx_rev is not None else 0
                mapped_data['판매원가'] = df.iloc[:, idx_cost] if idx_cost is not None else 0

                mapped_data['원본_NPV'] = df.iloc[:, idx_npv] if idx_npv is not None else 0
                mapped_data['원본_IRR'] = df.iloc[:, idx_irr] if idx_irr is not None else 0
                temp_clean_df = pd.DataFrame(mapped_data)
                clean_df_list.append(temp_clean_df)
            except Exception as e:
                pass

        if not clean_df_list:
            st.warning("분석 가능한 유효 데이터(기초자료)가 없습니다.")
        else:
            clean_df = pd.concat(clean_df_list, ignore_index=True)
            clean_df['구간명'] = clean_df['구간명'].astype(str).str.strip()

            clean_df['용도'] = clean_df['용도'].astype(str).str.strip()
            clean_df['용도'] = clean_df['용도'].replace('ROE', '투자보수율가산')
            clean_df['용도'] = clean_df['용도'].replace('연료전지', '연료전지용')

            valid_usages = ["공공택지", "공동주택", "산업용", "업무용", "영업용", "연료전지용", "주택용", "주택용(지자체)", "투자보수율가산"]
            clean_df = clean_df[clean_df['용도'].isin(valid_usages)]

            invalid_names = ['', '0', 'nan', 'none', 'null', '구간명', '소계', '합계', '총계', 'roe제외']
            clean_df = clean_df[~clean_df['구간명'].str.lower().isin(invalid_names)]
            clean_df = clean_df[~clean_df['구간명'].str.contains('합계|소계|총계|ROE', na=False, regex=True)]
            clean_df = clean_df.drop_duplicates(subset=['차수', '구간명'], keep='last')
            num_cols_base = ['길이', '투자비', '분담금', '기타이익', '총전수', '공동주택전수', '단독주택전수', '판매량', '판매액', '판매원가', '원본_NPV', '원본_IRR']
            for c in num_cols_base:
                if clean_df[c].dtype == object:
                    clean_df[c] = clean_df[c].astype(str).str.replace(',', '', regex=False)
                clean_df[c] = pd.to_numeric(clean_df[c], errors='coerce').fillna(0)
            clean_df['총전수'] = np.maximum(clean_df['총전수'], clean_df['공동주택전수'] + clean_df['단독주택전수'])

            clean_df['기본요금수익'] = 0.0
            temp_usage = clean_df['용도'].astype(str).str.replace(' ', '', regex=False)

            is_home = temp_usage.str.contains('주택|가정|공동|택지') & ~temp_usage.str.contains('외')
            clean_df.loc[is_home, '기본요금수익'] = clean_df.loc[is_home, '총전수'] * sim_basic_price * 12

            num_cols = ['길이', '투자비', '분담금', '기타이익', '총전수', '판매량', '판매액', '판매원가', '기본요금수익', '원본_NPV', '원본_IRR']
            st.success("✅ 기초자료 로드 완료! 아래에서 분석할 데이터의 범위를 선택해 주세요.")
            st.markdown("---")

            col1, col2 = st.columns(2)
            available_chas = sorted(clean_df['차수'].unique())

            with col1:
                selected_cha = st.selectbox("📌 기준 차수 선택", available_chas, index=len(available_chas)-1, format_func=lambda x: f"{x}차")
            with col2:
                view_mode = st.radio("보기 옵션 (데이터 조회 범위)", ["1. 당해차수 데이터", "2. 1차~현재까지 데이터"])
            if view_mode == "1. 당해차수 데이터":
                filtered_clean_df = clean_df[clean_df['차수'] == selected_cha]
                st.info(f"선택됨: **{selected_cha}차** 당해차수 데이터만 분석합니다.")
            else:
                filtered_clean_df = clean_df[clean_df['차수'] <= selected_cha]
                st.info(f"선택됨: **1차 부터 {selected_cha}차 까지의 누적** 데이터를 분석합니다.")

            st.markdown("---")
            def get_analysis_result(row, usage_val=''):
                npv_sim, irr_sim, irr_msg_sim, _ = calculate_simulation(
                    row['길이'], row['투자비'], row['분담금'], row['기타이익'], row['판매량'], row['판매액'], row['판매원가'],
                    row['총전수'], row['기본요금수익'], RATE, TAX, dep_period, analysis_period, c_maint, c_adm_jeon, c_adm_m
                )

                if '공공택지' in str(usage_val) or '주택용' in str(usage_val) or '영업용' in str(usage_val):
                    npv = row['원본_NPV']
                    irr = (row['원본_IRR'] / 100.0) if row['원본_IRR'] != 0 else 0
                    irr_msg = ""
                else:
                    npv = npv_sim
                    irr = irr_sim
                    irr_msg = irr_msg_sim

                return row['길이'], row['투자비'] - row['분담금'] - row['기타이익'], row['판매량'], npv, irr, irr_msg
            custom_order = ["공공택지", "공동주택", "산업용", "업무용", "영업용", "연료전지용", "주택용"]
            unique_usages = filtered_clean_df['용도'].unique().tolist()

            sorted_usages = sorted(unique_usages, key=lambda x: 9999 if x == '투자보수율가산' else (custom_order.index(x) if x in custom_order else 999))

            usage_results = []
            for u in sorted_usages:
                u_df = filtered_clean_df[filtered_clean_df['용도'] == u]
                u_len, u_net_inv, u_vol, u_npv, u_irr, u_irr_msg = get_analysis_result(u_df[num_cols].sum(), u)
                is_selected = False if '투자보수율가산' in str(u) else True

                u_inv_total = u_df['투자비'].sum()
                u_jeon_total = u_df['총전수'].sum()

                usage_results.append({
                    "선택": is_selected,
                    "용도": u,
                    "투자길이(m)": float(u_len),
                    "투자금액(원)": float(u_inv_total),
                    "공급전수(전)": float(u_jeon_total),
                    "연간판매량(MJ)": float(u_vol),
                    "NPV(원)": float(u_npv),
                    "IRR(%)": None if u == '투자보수율가산' else (float(u_irr*100) if u_irr is not None else None)
                })

            df_usage_summary = pd.DataFrame(usage_results)
            # session_state로 선택 상태 관리 (초기값 설정)
            ss_key = "usage_selection"
            if ss_key not in st.session_state:
                st.session_state[ss_key] = {row["용도"]: row["선택"] for row in usage_results}
            else:
                # 새 용도가 생기면 기본값 추가
                for row in usage_results:
                    if row["용도"] not in st.session_state[ss_key]:
                        st.session_state[ss_key][row["용도"]] = row["선택"]
            # session_state 기준으로 선택된 용도 결정
            selected_usages = [row["용도"] for row in usage_results if st.session_state[ss_key].get(row["용도"], row["선택"])]
            # ★ [1번] 선택 항목 합산 소계를 먼저 렌더링
            if selected_usages:
                final_filtered_df = filtered_clean_df[filtered_clean_df['용도'].isin(selected_usages)]
                t_len, t_net_inv, t_vol, tot_npv, tot_irr, tot_irr_msg = get_analysis_result(final_filtered_df[num_cols].sum(), '합산')

                tot_npv = sum(item["NPV(원)"] for item in usage_results if item["용도"] in selected_usages)
                st.subheader("1. 📊 선택 항목 합산 소계 (Subtotal)")
                m1, m2 = st.columns(2)
                m1.metric("최종 합산 NPV", f"{tot_npv:,.0f} 원")
                m2.metric("최종 합산 IRR", f"{tot_irr*100:.2f} %" if tot_irr is not None else tot_irr_msg)

                t_inv_total = final_filtered_df['투자비'].sum()
                t_jeon_total = final_filtered_df['총전수'].sum()

                subtotal_df = pd.DataFrame([{
                    "항목명": "☑️ 선택 용도 총합계",
                    "투자길이(m)": t_len, "투자금액(원)": t_inv_total, "공급전수(전)": t_jeon_total, "연간판매량(MJ)": t_vol, "NPV(원)": tot_npv
                }])

                st.dataframe(subtotal_df.style.format({
                    "투자길이(m)": "{:,.0f}",
                    "투자금액(원)": "{:,.0f}",
                    "공급전수(전)": "{:,.0f}",
                    "연간판매량(MJ)": "{:,.0f}",
                    "NPV(원)": "{:,.0f}"
                }), hide_index=True)
                st.divider()
            # ★ [2번] 용도별 경제성 요약 — data_editor를 여기에만 렌더링
            st.subheader("2. 📁 용도별 경제성 요약 (분석 대상 선택)")
            # session_state 선택값을 df에 반영
            df_usage_for_editor = df_usage_summary.copy()
            df_usage_for_editor["선택"] = df_usage_for_editor["용도"].map(
                lambda u: st.session_state[ss_key].get(u, True)
            )
            edited_df = st.data_editor(
                df_usage_for_editor.style.format({
                    "투자길이(m)": "{:,.0f}",
                    "투자금액(원)": "{:,.0f}",
                    "공급전수(전)": "{:,.0f}",
                    "연간판매량(MJ)": "{:,.0f}",
                    "NPV(원)": "{:,.0f}",
                    "IRR(%)": lambda x: f"{x:,.2f}" if pd.notnull(x) else ""
                }),
                column_config={
                    "선택": st.column_config.CheckboxColumn("선택")
                },
                disabled=["용도", "투자길이(m)", "투자금액(원)", "공급전수(전)", "연간판매량(MJ)", "NPV(원)", "IRR(%)"],
                hide_index=True,
                use_container_width=True,
                key="usage_editor_main"
            )
            # 체크박스 변경 시 session_state 업데이트 후 재실행
            new_selection = {row["용도"]: row["선택"] for _, row in edited_df.iterrows()}
            if new_selection != st.session_state[ss_key]:
                st.session_state[ss_key] = new_selection
                st.rerun()
            if selected_usages:
                st.divider()
                st.subheader("3. 📑 구간별 경제성 상세 명세서")
                df_detail = final_filtered_df.groupby(['용도', '구간명'])[num_cols].sum().reset_index()

                df_detail['용도_순위'] = df_detail['용도'].apply(lambda x: 9999 if x == '투자보수율가산' else (custom_order.index(x) if x in custom_order else 999))
                df_detail = df_detail.sort_values(by=['용도_순위', '구간명']).drop(columns=['용도_순위'])

                detail_results = []
                for _, row in df_detail.iterrows():
                    d_len, d_net_inv, d_vol, d_npv, d_irr, d_irr_msg = get_analysis_result(row, row['용도'])

                    detail_results.append({
                        "용도": row['용도'], "구간명": row['구간명'],
                        "투자길이(m)": d_len, "투자금액(원)": row['투자비'],
                        "공급전수(전)": row['총전수'], "연간판매량(MJ)": d_vol,
                        "NPV(원)": d_npv, "IRR(%)": None if row['용도'] == '투자보수율가산' else (d_irr * 100 if d_irr is not None else None)
                    })

                if detail_results:
                    sub_row = {
                        "용도": "소계",
                        "구간명": "",
                        "투자길이(m)": sum(x["투자길이(m)"] for x in detail_results),
                        "투자금액(원)": sum(x["투자금액(원)"] for x in detail_results),
                        "공급전수(전)": sum(x["공급전수(전)"] for x in detail_results),
                        "연간판매량(MJ)": sum(x["연간판매량(MJ)"] for x in detail_results),
                        "NPV(원)": sum(x["NPV(원)"] for x in detail_results),
                        "IRR(%)": None
                    }
                    detail_results.append(sub_row)

                df_detail_final = pd.DataFrame(detail_results)

                def highlight_subtotal(row):
                    if row['용도'] == '소계':
                        return ['background-color: #E0E0E0; font-weight: bold'] * len(row)
                    return [''] * len(row)

                st.dataframe(df_detail_final.style.apply(highlight_subtotal, axis=1).format({
                    "투자길이(m)": "{:,.0f}", "투자금액(원)": "{:,.0f}", "공급전수(전)": "{:,.0f}",
                    "연간판매량(MJ)": "{:,.0f}", "NPV(원)": "{:,.0f}",
                    "IRR(%)": lambda x: f"{x:,.2f}" if pd.notnull(x) else ""
                }), use_container_width=True, hide_index=True)
    else:
        st.info("👆 좌측 사이드바에 파일이 로드되지 않았습니다. 깃허브에 파일을 올리거나 직접 업로드 해주세요.")
# ==========================================================================
# 탭 2: 신규 배관 투자 승인 내역 자동화
# ==========================================================================
elif menu_choice == '2. 배관 투자 승인 내역':

    extracted_bp_data_by_cha = {}
    all_data_unfiltered = []
    chas = []

    def get_multi_col_idx(df_temp, keywords):
        found_cols = []
        for col_idx in range(df_temp.shape[1]):
            for row_idx in range(min(20, df_temp.shape[0])):
                val = str(df_temp.iloc[row_idx, col_idx]).replace(" ", "").replace("\n", "")
                for kw in keywords:
                    if kw in val:
                        found_cols.append(col_idx)
        return list(set(found_cols))

    if working_files:
        for file_obj in working_files:
            try:
                file_name = file_obj.name if hasattr(file_obj, 'name') else os.path.basename(file_obj)
                if hasattr(file_obj, 'seek'):
                    file_obj.seek(0)

                # 품의서 파일은 탭2에서 제외
                if '품의서' in file_name:
                    continue

                match = re.search(r'(\d+)차', file_name)
                file_cha = int(match.group(1)) if match else 1
                chas.append(file_cha)

                # 승인내역 파일은 탭2에서도 제외 (저장용 파일)
                if '승인내역' in file_name:
                    continue

                # A. 공무팀 현황정리 파일
                if '현황' in file_name or ('투자' in file_name and '기초자료' not in file_name):
                    if file_name.endswith('.csv'):
                        df_status = pd.read_csv(file_obj, header=None, encoding='utf-8-sig')
                    else:
                        try:
                            df_status = pd.read_excel(file_obj, sheet_name='투자현황', header=None)
                        except:
                            df_status = pd.read_excel(file_obj, header=None)
                    if file_cha not in extracted_bp_data_by_cha:
                        extracted_bp_data_by_cha[file_cha] = {}

                    for row_idx in range(df_status.shape[0]):
                        col1_val = str(df_status.iloc[row_idx, 1]).strip()
                        clean_col1 = col1_val.replace('\n', '').replace(' ', '')
                        target_items = ['계획배관', 'Loop', '이설배관', '지역정압기', '공급시설물개선', '인입배관']

                        for item in target_items:
                            if item in clean_col1:
                                try:
                                    p_scale = pd.to_numeric(str(df_status.iloc[row_idx, 12]).replace(',', ''), errors='coerce')
                                    p_amt = pd.to_numeric(str(df_status.iloc[row_idx, 13]).replace(',', ''), errors='coerce')
                                    c_scale = pd.to_numeric(str(df_status.iloc[row_idx, 15]).replace(',', ''), errors='coerce')
                                    c_amt = pd.to_numeric(str(df_status.iloc[row_idx, 16]).replace(',', ''), errors='coerce')

                                    extracted_bp_data_by_cha[file_cha][item] = {
                                        '기승인_규모': p_scale if not pd.isna(p_scale) else 0,
                                        '기승인_금액': p_amt if not pd.isna(p_amt) else 0,
                                        '금회_규모': c_scale if not pd.isna(c_scale) else 0,
                                        '금회_금액': c_amt if not pd.isna(c_amt) else 0
                                    }
                                except:
                                    pass

                # B. 수요개발배관(기초자료) 파일
                else:
                    if file_name.endswith('.csv'):
                        df = pd.read_csv(file_obj, header=None, encoding='utf-8-sig')
                    else:
                        df = pd.read_excel(file_obj, header=None)
                    df = df.iloc[:100]
                    extracted = pd.DataFrame()

                    extracted['항목'] = df.iloc[:, 0].astype(str).str.strip()
                    extracted['항목'] = extracted['항목'].replace('ROE', '투자보수율가산')
                    extracted['항목'] = extracted['항목'].replace('연료전지', '연료전지용')

                    idx_name = get_col_idx(df, ["구간명"], exact=True)
                    idx_len = get_col_idx(df, ["길이(m)", "배관길이"], exact=False)
                    idx_inv = get_col_idx(df, ["배관투자금액", "총공사비"], exact=False)
                    idx_jeon = get_col_idx(df, ["수요전수계", "총전수"], exact=False)
                    idx_jeon_apt = get_col_idx(df, ["공동주택전수"], exact=False)
                    idx_jeon_single = get_col_idx(df, ["단독주택전수"], exact=False)
                    idx_total_vol = get_col_idx(df, ["계(MJ)"], exact=False)
                    idx_npv = get_col_idx(df, ["NPV"], exact=False)
                    idx_irr = get_col_idx(df, ["IRR"], exact=False)
                    def get_clean_series(c_idx):
                        return pd.to_numeric(df.iloc[:, c_idx].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
                    extracted['규모(m)'] = get_clean_series(idx_len) if idx_len is not None else 0
                    extracted['투자비(원)'] = get_clean_series(idx_inv) if idx_inv is not None else 0

                    extracted['공사명'] = df.iloc[:, idx_name] if idx_name is not None else df.iloc[:, 1]
                    extracted['차수'] = file_cha

                    jeon_total = get_clean_series(idx_jeon) if idx_jeon is not None else pd.Series(0, index=df.index)
                    jeon_apt = get_clean_series(idx_jeon_apt) if idx_jeon_apt is not None else pd.Series(0, index=df.index)
                    jeon_single = get_clean_series(idx_jeon_single) if idx_jeon_single is not None else pd.Series(0, index=df.index)
                    extracted['전수(전)'] = np.maximum(jeon_total, jeon_apt + jeon_single)
                    extracted['가정용 판매량(MJ)'] = 0
                    extracted['일반용 판매량(MJ)'] = 0
                    extracted['판매량(MJ)'] = get_clean_series(idx_total_vol) if idx_total_vol is not None else 0
                    extracted['NPV(원)'] = get_clean_series(idx_npv) if idx_npv is not None else 0
                    extracted['IRR(%)'] = get_clean_series(idx_irr) if idx_irr is not None else 0
                    valid_target_usages = ["공공택지", "공동주택", "산업용", "업무용", "영업용", "연료전지용", "주택용", "주택용(지자체)", "투자보수율가산"]
                    extracted = extracted[extracted['항목'].isin(valid_target_usages)]

                    extracted['공사명'] = extracted['공사명'].astype(str).str.strip()
                    extracted = extracted[~extracted['공사명'].str.contains('합계|소계|총계|roe|제외|구간명', case=False, na=False, regex=True)]

                    def clean_numeric(x):
                        s = str(x).replace(',', '')
                        s = re.sub(r'[^\d.-]', '', s)
                        try:
                            return float(s) if s else 0.0
                        except:
                            return 0.0
                    num_cols_ext = ['규모(m)', '투자비(원)', '전수(전)', '판매량(MJ)', 'NPV(원)', 'IRR(%)']
                    for c in num_cols_ext:
                        extracted[c] = extracted[c].apply(clean_numeric)

                    all_data_unfiltered.append(extracted)
            except Exception as e:
                pass
    # --- 메인 화면 상단 ---
    st.title("📋 2026년도 배관 투자 승인 내역")
    st.markdown("깃허브 파일 데이터를 바탕으로 **수요개발은 자동 계산**하고, **기본계획/인입은 공무팀 실적 파일과 연동**하여 산출합니다.")
    if working_files:
        available_chas_t2 = sorted(list(set(chas))) if chas else [1]

        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            selected_cha_t2 = st.selectbox("📌 기준 차수 선택 (승인 내역용)", available_chas_t2, index=len(available_chas_t2)-1, format_func=lambda x: f"{x}차")
        with col2:
            view_mode_t2 = st.radio("보기 옵션 (데이터 조회 범위)", ["1. 당해차수 데이터", "2. 1차~현재까지 데이터"], key="view_mode_t2")

        st.markdown("---")
    else:
        selected_cha_t2 = 1
        view_mode_t2 = "1. 당해차수 데이터"
    # --- 좌측 사이드바 ---
    with st.sidebar:
        st.header("💰 2026년 사업계획 투자한도액 세팅")

        st.subheader("🔹 1. 수요개발배관 (실적 자동 추출)")
        df_sd_base = pd.DataFrame({
            "항목": ["공공택지", "공동주택", "산업용", "업무용", "영업용", "연료전지용", "주택용", "주택용(지자체)", "투자보수율가산"],
            "한도_규모": [988, 794, 46, 476, 180, 505, 16, 0, 1972],
            "한도_금액": [825459960, 1032719000, 81162000, 439210000, 119752000, 503524000, 20205000, 0, 1646562000]
        })
        edited_sd = st.data_editor(df_sd_base, key="sd_editor", hide_index=True, use_container_width=True)

        st.divider()
        available_bp_chas = [c for c in extracted_bp_data_by_cha.keys() if c <= selected_cha_t2]
        if available_bp_chas:
            latest_bp_cha = max(available_bp_chas)
            latest_bp_data = extracted_bp_data_by_cha[latest_bp_cha]
        else:
            latest_bp_cha = 0
            latest_bp_data = {}
        st.subheader("🔹 2. 기본계획배관 (현황정리 엑셀 연동)")
        if latest_bp_cha == selected_cha_t2:
            st.markdown(f"**✅ 현재 선택된 '{selected_cha_t2}차' 현황정리가 적용되었습니다.**")
        elif latest_bp_cha > 0:
            st.markdown(f"**⚠️ '{selected_cha_t2}차' 파일이 없어, 가장 최근인 '{latest_bp_cha}차'의 실적 누계가 '기승인'으로 자동 이월되었습니다!**")
        else:
            st.markdown(f"**⚠️ 해당하는 현황정리 데이터가 없습니다.**")

        bp_items = ["계획배관", "Loop", "이설배관", "지역정압기", "공급시설물 개선"]
        bp_limit_s = [2580, 155, 869, 3, 95]
        bp_limit_a = [1317338580, 158921000, 1790515000, 348873000, 2749999000]

        bp_p_s, bp_p_a, bp_c_s, bp_c_a = [], [], [], []

        for item in bp_items:
            clean_item = item.replace('\n', '').replace(' ', '')
            if clean_item in latest_bp_data:
                if latest_bp_cha == selected_cha_t2:
                    bp_p_s.append(latest_bp_data[clean_item]['기승인_규모'])
                    bp_p_a.append(latest_bp_data[clean_item]['기승인_금액'])
                    bp_c_s.append(latest_bp_data[clean_item]['금회_규모'])
                    bp_c_a.append(latest_bp_data[clean_item]['금회_금액'])
                else:
                    bp_p_s.append(latest_bp_data[clean_item]['기승인_규모'] + latest_bp_data[clean_item]['금회_규모'])
                    bp_p_a.append(latest_bp_data[clean_item]['기승인_금액'] + latest_bp_data[clean_item]['금회_금액'])
                    bp_c_s.append(0)
                    bp_c_a.append(0)
            else:
                bp_p_s.append(0); bp_p_a.append(0); bp_c_s.append(0); bp_c_a.append(0)
        df_bp_base = pd.DataFrame({
            "항목": bp_items,
            "한도_규모": bp_limit_s,
            "한도_금액": bp_limit_a,
            "기승인_규모": bp_p_s,
            "기승인_금액": bp_p_a,
            "금회_규모": bp_c_s,
            "금회_금액": bp_c_a
        })
        edited_bp = st.data_editor(df_bp_base, key=f"bp_editor_{selected_cha_t2}", hide_index=True, use_container_width=True)
        st.divider()
        st.subheader("🔹 3. 65A미만 인입 (현황정리 엑셀 연동)")

        in_p_s, in_p_a, in_c_s, in_c_a = 0, 0, 0, 0
        clean_item = '인입배관'
        if clean_item in latest_bp_data:
            if latest_bp_cha == selected_cha_t2:
                in_p_s = latest_bp_data[clean_item]['기승인_규모']
                in_p_a = latest_bp_data[clean_item]['기승인_금액']
                in_c_s = latest_bp_data[clean_item]['금회_규모']
                in_c_a = latest_bp_data[clean_item]['금회_금액']
            else:
                in_p_s = latest_bp_data[clean_item]['기승인_규모'] + latest_bp_data[clean_item]['금회_규모']
                in_p_a = latest_bp_data[clean_item]['기승인_금액'] + latest_bp_data[clean_item]['금회_금액']
                in_c_s = 0
                in_c_a = 0
        df_in_base = pd.DataFrame({
            "항목": ["65A미만 인입"],
            "한도_규모": [730],
            "한도_금액": [3231095000],
            "기승인_규모": [in_p_s],
            "기승인_금액": [in_p_a],
            "금회_규모": [in_c_s],
            "금회_금액": [in_c_a]
        })
        edited_in = st.data_editor(df_in_base, key=f"in_editor_{selected_cha_t2}", hide_index=True, use_container_width=True)
    # --- 메인 화면 하단: 표 및 그래프 렌더링 ---
    if working_files:
        if all_data_unfiltered:
            all_parsed_df = pd.concat(all_data_unfiltered, ignore_index=True)

            all_parsed_df = all_parsed_df.drop_duplicates(subset=['차수', '항목', '공사명'], keep='last')
            all_parsed_df['항목_clean'] = all_parsed_df['항목'].astype(str).str.replace(r'\s+', '', regex=True)

            prev_df = all_parsed_df[all_parsed_df['차수'] < selected_cha_t2]
            curr_df = all_parsed_df[all_parsed_df['차수'] == selected_cha_t2]

            prev_agg = prev_df.groupby('항목_clean')[['규모(m)', '투자비(원)']].sum()
            curr_agg = curr_df.groupby('항목_clean')[['규모(m)', '투자비(원)']].sum()
        else:
            all_parsed_df = pd.DataFrame()
            prev_agg = pd.DataFrame()
            curr_agg = pd.DataFrame()
        def fill_sd_metrics(df_base):
            for col in ['기승인_규모', '기승인_금액', '금회_규모', '금회_금액', '누계_규모', '누계_금액']:
                df_base[col] = 0.0

            for i in range(len(df_base)):
                item_raw = df_base.at[i, '항목']
                if item_raw == '소계': continue

                item_clean = str(item_raw).replace('\n', '').replace(' ', '')

                if not prev_agg.empty:
                    matched_idx = [idx for idx in prev_agg.index if item_clean == str(idx)]
                    if matched_idx:
                        df_base.at[i, '기승인_규모'] = prev_agg.loc[matched_idx, '규모(m)'].sum()
                        df_base.at[i, '기승인_금액'] = prev_agg.loc[matched_idx, '투자비(원)'].sum()

                if not curr_agg.empty:
                    matched_idx = [idx for idx in curr_agg.index if item_clean == str(idx)]
                    if matched_idx:
                        df_base.at[i, '금회_규모'] = curr_agg.loc[matched_idx, '규모(m)'].sum()
                        df_base.at[i, '금회_금액'] = curr_agg.loc[matched_idx, '투자비(원)'].sum()

                df_base.at[i, '누계_규모'] = df_base.at[i, '기승인_규모'] + df_base.at[i, '금회_규모']
                df_base.at[i, '누계_금액'] = df_base.at[i, '기승인_금액'] + df_base.at[i, '금회_금액']
                df_base.at[i, '잔여_금액'] = df_base.at[i, '한도_금액'] - df_base.at[i, '누계_금액']

            sub_idx = len(df_base) - 1
            for col in ['기승인_규모', '기승인_금액', '금회_규모', '금회_금액', '누계_규모', '누계_금액', '잔여_금액']:
                df_base.at[sub_idx, col] = df_base.iloc[:-1][col].sum()

            return df_base
        df_sd_display = edited_sd.copy()
        df_sd_display.rename(columns={'규모': '한도_규모', '금액': '한도_금액'}, inplace=True)
        df_sd_display.insert(0, '구분', '수요개발배관')
        df_sd_display.loc[len(df_sd_display)] = ['수요개발배관', '소계', df_sd_display['한도_규모'].sum(), df_sd_display['한도_금액'].sum()]
        df_sd_display = fill_sd_metrics(df_sd_display)

        df_bp_display = edited_bp.copy()
        df_bp_display.insert(0, '구분', '기본계획배관')
        df_bp_display['누계_규모'] = df_bp_display['기승인_규모'] + df_bp_display['금회_규모']
        df_bp_display['누계_금액'] = df_bp_display['기승인_금액'] + df_bp_display['금회_금액']
        df_bp_display['잔여_금액'] = df_bp_display['한도_금액'] - df_bp_display['누계_금액']
        bp_sub_row = {
            '구분': '기본계획배관', '항목': '소계',
            '한도_규모': df_bp_display['한도_규모'].sum(), '한도_금액': df_bp_display['한도_금액'].sum(),
            '기승인_규모': df_bp_display['기승인_규모'].sum(), '기승인_금액': df_bp_display['기승인_금액'].sum(),
            '금회_규모': df_bp_display['금회_규모'].sum(), '금회_금액': df_bp_display['금회_금액'].sum(),
            '누계_규모': df_bp_display['누계_규모'].sum(), '누계_금액': df_bp_display['누계_금액'].sum(),
            '잔여_금액': df_bp_display['잔여_금액'].sum()
        }
        df_bp_display = pd.concat([df_bp_display, pd.DataFrame([bp_sub_row])], ignore_index=True)
        df_in_display = edited_in.copy()
        df_in_display.insert(0, '구분', '65A미만 인입')
        df_in_display['누계_규모'] = df_in_display['기승인_규모'] + df_in_display['금회_규모']
        df_in_display['누계_금액'] = df_in_display['기승인_금액'] + df_in_display['금회_금액']
        df_in_display['잔여_금액'] = df_in_display['한도_금액'] - df_in_display['누계_금액']
        in_sub_row = {
            '구분': '65A미만 인입', '항목': '소계',
            '한도_규모': df_in_display['한도_규모'].sum(), '한도_금액': df_in_display['한도_금액'].sum(),
            '기승인_규모': df_in_display['기승인_규모'].sum(), '기승인_금액': df_in_display['기승인_금액'].sum(),
            '금회_규모': df_in_display['금회_규모'].sum(), '금회_금액': df_in_display['금회_금액'].sum(),
            '누계_규모': df_in_display['누계_규모'].sum(), '누계_금액': df_in_display['누계_금액'].sum(),
            '잔여_금액': df_in_display['잔여_금액'].sum()
        }
        df_in_display = pd.concat([df_in_display, pd.DataFrame([in_sub_row])], ignore_index=True)

        df_budget_detail = pd.concat([df_sd_display, df_bp_display, df_in_display], ignore_index=True)

        sd_sub_idx = len(df_sd_display) - 1
        bp_sub_idx = len(df_bp_display) - 1
        in_sub_idx = len(df_in_display) - 1

        tot_limit_scale = df_sd_display.at[sd_sub_idx, '한도_규모'] + df_bp_display.at[bp_sub_idx, '한도_규모'] + df_in_display.at[in_sub_idx, '한도_규모']
        tot_limit_amt = df_sd_display.at[sd_sub_idx, '한도_금액'] + df_bp_display.at[bp_sub_idx, '한도_금액'] + df_in_display.at[in_sub_idx, '한도_금액']
        tot_prev_scale = df_sd_display.at[sd_sub_idx, '기승인_규모'] + df_bp_display.at[bp_sub_idx, '기승인_규모'] + df_in_display.at[in_sub_idx, '기승인_규모']
        tot_prev_amt = df_sd_display.at[sd_sub_idx, '기승인_금액'] + df_bp_display.at[bp_sub_idx, '기승인_금액'] + df_in_display.at[in_sub_idx, '기승인_금액']
        tot_curr_scale = df_sd_display.at[sd_sub_idx, '금회_규모'] + df_bp_display.at[bp_sub_idx, '금회_규모'] + df_in_display.at[in_sub_idx, '금회_규모']
        tot_curr_amt = df_sd_display.at[sd_sub_idx, '금회_금액'] + df_bp_display.at[bp_sub_idx, '금회_금액'] + df_in_display.at[in_sub_idx, '금회_금액']
        tot_cum_scale = df_sd_display.at[sd_sub_idx, '누계_규모'] + df_bp_display.at[bp_sub_idx, '누계_규모'] + df_in_display.at[in_sub_idx, '누계_규모']
        tot_cum_amt = df_sd_display.at[sd_sub_idx, '누계_금액'] + df_bp_display.at[bp_sub_idx, '누계_금액'] + df_in_display.at[in_sub_idx, '누계_금액']
        tot_remain_amt = df_sd_display.at[sd_sub_idx, '잔여_금액'] + df_bp_display.at[bp_sub_idx, '잔여_금액'] + df_in_display.at[in_sub_idx, '잔여_금액']
        df_budget_detail.loc[len(df_budget_detail)] = [
            '합계', '총계',
            tot_limit_scale, tot_limit_amt,
            tot_prev_scale, tot_prev_amt,
            tot_curr_scale, tot_curr_amt,
            tot_cum_scale, tot_cum_amt,
            tot_remain_amt
        ]
        st.subheader("📈 2026년 배관 투자 전체 요약 및 진척도")

        sum_df = pd.DataFrame({
            "구분": ["수요개발배관", "기본계획배관", "65A미만 인입", "총계"],
            "총 투자한도액": [df_sd_display.at[sd_sub_idx, '한도_금액'], df_bp_display.at[bp_sub_idx, '한도_금액'], df_in_display.at[in_sub_idx, '한도_금액'], tot_limit_amt],
            "기승인 누계": [df_sd_display.at[sd_sub_idx, '기승인_금액'], df_bp_display.at[bp_sub_idx, '기승인_금액'], df_in_display.at[in_sub_idx, '기승인_금액'], tot_prev_amt],
            f"금회 신청 ({selected_cha_t2}차)": [df_sd_display.at[sd_sub_idx, '금회_금액'], df_bp_display.at[bp_sub_idx, '금회_금액'], df_in_display.at[in_sub_idx, '금회_금액'], tot_curr_amt],
            "현재 본부 누계": [df_sd_display.at[sd_sub_idx, '누계_금액'], df_bp_display.at[bp_sub_idx, '누계_금액'], df_in_display.at[in_sub_idx, '누계_금액'], tot_cum_amt],
            "잔여 한도액": [df_sd_display.at[sd_sub_idx, '잔여_금액'], df_bp_display.at[bp_sub_idx, '잔여_금액'], df_in_display.at[in_sub_idx, '잔여_금액'], tot_remain_amt]
        })
        sum_df['집행률(%)'] = np.where(sum_df['총 투자한도액'] > 0, (sum_df['현재 본부 누계'] / sum_df['총 투자한도액']) * 100, 0)

        st.dataframe(sum_df.style.format({
            "총 투자한도액": "{:,.0f}", "기승인 누계": "{:,.0f}", f"금회 신청 ({selected_cha_t2}차)": "{:,.0f}",
            "현재 본부 누계": "{:,.0f}", "잔여 한도액": "{:,.0f}", "집행률(%)": "{:,.1f}%"
        }).apply(lambda x: ['background-color: #FFE6E6; font-weight: bold'] * len(x) if x['구분'] == '총계' else [''] * len(x), axis=1),
        hide_index=True, use_container_width=True)
        st.markdown("<br>", unsafe_allow_html=True)

        col_c1, col_c2 = st.columns([2, 1.5])

        with col_c1:
            st.markdown("#### 📊 구분별 투자 예산 및 실적 현황")
            chart_df = sum_df.iloc[:3].copy()
            bar_data = pd.DataFrame({
                "구분": chart_df["구분"].tolist() * 2,
                "금액(원)": chart_df["총 투자한도액"].tolist() + chart_df["현재 본부 누계"].tolist(),
                "유형": ["총 투자한도액"] * 3 + ["현재 본부 누계"] * 3
            })
            fig_bar = px.bar(bar_data, x="구분", y="금액(원)", color="유형", barmode="group",
                             color_discrete_map={"총 투자한도액": "#1E88E5", "현재 본부 누계": "#FFB300"})
            fig_bar.update_layout(margin=dict(t=20, b=0, l=0, r=0), legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
            st.plotly_chart(fig_bar, use_container_width=True)
        with col_c2:
            st.markdown("#### 🍩 전체 투자계획 대비 실적")
            donut_rem = tot_remain_amt if tot_remain_amt > 0 else 0
            donut_df = pd.DataFrame({
                "상태": ["실적 (누계)", "잔여 한도액"],
                "금액": [tot_cum_amt, donut_rem]
            })
            fig_donut = px.pie(donut_df, values='금액', names='상태', hole=0.5,
                               color='상태', color_discrete_map={"실적 (누계)": "#FFB300", "잔여 한도액": "#E0E0E0"})
            fig_donut.update_traces(textinfo='percent+label', textposition='inside')

            exec_rate = (tot_cum_amt / tot_limit_amt * 100) if tot_limit_amt > 0 else 0
            fig_donut.update_layout(showlegend=False, margin=dict(t=20, b=0, l=0, r=0),
                                    annotations=[dict(text=f"집행률<br><b>{exec_rate:.1f}%</b>", x=0.5, y=0.5, font_size=18, showarrow=False)])
            st.plotly_chart(fig_donut, use_container_width=True)
        st.divider()
        st.subheader("📌 2026년 배관 투자 승인 요약 (Excel 양식)")
        df_budget_detail['승인비율'] = np.where(df_budget_detail['한도_금액'] > 0, (df_budget_detail['누계_금액'] / df_budget_detail['한도_금액']) * 100, 0)
        df_budget_detail['구분'] = df_budget_detail['구분'].mask(df_budget_detail['구분'].duplicated(), '')
        df_budget_detail.columns = pd.MultiIndex.from_tuples([
            ('구분', ''), ('항목', ''),
            ('2026년 사업계획 투자한도액', '규모(m)'), ('2026년 사업계획 투자한도액', '금액'),
            ('2026년 본부투자 승인내역 (기승인)', '규모(m)'), ('2026년 본부투자 승인내역 (기승인)', '금액'),
            (f'금회 신청 내역 ({selected_cha_t2}차)', '규모(m)'), (f'금회 신청 내역 ({selected_cha_t2}차)', '금액'),
            ('2026년 본부투자 누계', '규모(m)'), ('2026년 본부투자 누계', '금액'),
            ('투자한도 잔액', '금액'), ('승인비율', '(%)')
        ])
        def style_rows(row):
            if row[('항목', '')] == '소계':
                return ['background-color: #E6F3FF; font-weight: bold'] * len(row)
            elif row[('항목', '')] == '총계':
                return ['background-color: #FFE6E6; font-weight: bold; color: #D32F2F'] * len(row)
            return [''] * len(row)
        format_dict = {col: "{:,.0f}" for col in df_budget_detail.columns if col[0] not in ['구분', '항목', '승인비율']}
        format_dict[('승인비율', '(%)')] = "{:,.1f}%"
        styled_df = df_budget_detail.style.format(format_dict).apply(style_rows, axis=1)
        st.dataframe(styled_df, hide_index=True, use_container_width=True)
        st.divider()

        if not all_parsed_df.empty:
            if view_mode_t2 == "1. 당해차수 데이터":
                detail_df = all_parsed_df[all_parsed_df['차수'] == selected_cha_t2]
            else:
                detail_df = all_parsed_df[all_parsed_df['차수'] <= selected_cha_t2]

            if not detail_df.empty:
                detail_title_prefix = f"{selected_cha_t2}차 당해" if view_mode_t2 == "1. 당해차수 데이터" else f"{selected_cha_t2}차 누계"

                detail_df = detail_df[~detail_df['항목'].astype(str).str.lower().str.contains('none|nan|null|^0$|^$', regex=True, na=False)]
                detail_df = detail_df[(detail_df['규모(m)'] != 0) | (detail_df['투자비(원)'] != 0) | (detail_df['판매량(MJ)'] != 0)]

                st.subheader(f"📊 {detail_title_prefix} 용도별 실적 요약")

                usage_counts = detail_df.groupby('항목').size().reset_index(name='건수')

                usage_sums = detail_df.groupby('항목').agg({
                    '규모(m)': 'sum', '투자비(원)': 'sum', '전수(전)': 'sum',
                    '판매량(MJ)': 'sum', 'NPV(원)': 'sum'
                }).reset_index()

                usage_summary = pd.merge(usage_counts, usage_sums, on='항목')

                valid_irr_df = detail_df[detail_df['항목'] != '투자보수율가산']
                irr_means = valid_irr_df.groupby('항목')['IRR(%)'].mean().reset_index()
                usage_summary = pd.merge(usage_summary, irr_means, on='항목', how='left')

                custom_order = ["공공택지", "공동주택", "산업용", "업무용", "영업용", "연료전지용", "주택용"]
                usage_summary['용도_순위'] = usage_summary['항목'].apply(lambda x: 9999 if x == '투자보수율가산' else (custom_order.index(x) if x in custom_order else 999))
                usage_summary = usage_summary.sort_values(by=['용도_순위']).drop(columns=['용도_순위'])

                total_irr = valid_irr_df['IRR(%)'].mean() if not valid_irr_df.empty else None

                usage_summary.loc[len(usage_summary)] = [
                    '총계',
                    usage_summary['건수'].sum(),
                    usage_summary['규모(m)'].sum(),
                    usage_summary['투자비(원)'].sum(),
                    usage_summary['전수(전)'].sum(),
                    usage_summary['판매량(MJ)'].sum(),
                    usage_summary['NPV(원)'].sum(),
                    total_irr
                ]

                st.dataframe(usage_summary.style.format({
                    "건수": "{:,.0f} 건",
                    "규모(m)": "{:,.0f}",
                    "투자비(원)": "{:,.0f}",
                    "전수(전)": "{:,.0f}",
                    "판매량(MJ)": "{:,.0f}",
                    "NPV(원)": "{:,.0f}",
                    "IRR(%)": lambda x: f"{x:,.2f}" if pd.notnull(x) else ""
                }).apply(lambda x: ['background-color: #F5F5F5; font-weight: bold'] * len(x) if x['항목'] == '총계' else [''] * len(x), axis=1),
                hide_index=True, use_container_width=True)

                st.markdown("<br>", unsafe_allow_html=True)

                st.subheader(f"📝 {detail_title_prefix} 상세 승인 내역 리스트")

                detail_df = detail_df.reset_index(drop=True)

                detail_df['용도_순위'] = detail_df['항목'].apply(lambda x: 9999 if x == '투자보수율가산' else (custom_order.index(x) if x in custom_order else 999))
                detail_df = detail_df.sort_values(by=['용도_순위', '공사명']).drop(columns=['용도_순위']).reset_index(drop=True)

                detail_df.loc[detail_df['항목'] == '투자보수율가산', 'IRR(%)'] = np.nan

                display_cols = ['항목', '공사명', '규모(m)', '투자비(원)', '전수(전)', '판매량(MJ)', 'NPV(원)', 'IRR(%)']

                st.dataframe(detail_df[display_cols].style.format({
                    "규모(m)": "{:,.0f}",
                    "투자비(원)": "{:,.0f}",
                    "전수(전)": "{:,.0f}",
                    "판매량(MJ)": "{:,.0f}",
                    "NPV(원)": "{:,.0f}",
                    "IRR(%)": lambda x: f"{x:,.2f}" if pd.notnull(x) else ""
                }), use_container_width=True, hide_index=True)

                csv_data = detail_df[display_cols].to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    label="📥 상세 승인내역 CSV 다운로드",
                    data=csv_data,
                    file_name=f"2026년도_배관투자_승인내역_{detail_title_prefix}.csv",
                    mime="text/csv"
                )
            else:
                st.info("해당 차수에 추출된 세부 공사 내역이 없습니다.")
        else:
            st.info("조건에 맞는 데이터가 없습니다.")


# ==========================================================================
# 탭 3: 품의서 결재 (기초자료에서 자동 생성)
# ==========================================================================
elif menu_choice == '3. 품의서 결재':

    # ── 사이드바: 분석 변수 ──
    with st.sidebar:
        st.header("⚙️ 분석 변수 설정")
        rate_pct_t3 = st.number_input("할인율 (%)", value=5.86, step=0.01, format="%.2f", key="t3_rate")
        tax_pct_t3 = st.number_input("법인세율+주민세율 (%)", value=22.0, step=0.1, format="%.1f", key="t3_tax")
        dep_period_t3 = st.number_input("감가상각 연수 (년)", value=30, step=1, key="t3_dep")
        analysis_period_t3 = st.number_input("경제성 분석 연수 (년)", value=30, step=1, key="t3_ana")
        st.subheader("💰 비용 단가")
        c_maint_t3 = st.number_input("유지비 (원/m)", value=8337, format="%d", key="t3_maint")
        c_adm_jeon_t3 = st.number_input("관리비 (원/전)", value=17357, format="%d", key="t3_adm_jeon")
        c_adm_m_t3 = st.number_input("관리비 (원/m)", value=11870, format="%d", key="t3_adm_m")
        sim_basic_price_t3 = st.number_input("주택용 월 기본요금 단가 (원)", value=900, step=10, format="%d", key="t3_basic")
        RATE_T3 = rate_pct_t3 / 100
        TAX_T3 = tax_pct_t3 / 100

    st.title("📄 품의서 결재")
    st.markdown("깃허브(또는 업로드)된 **기초자료**를 자동으로 파싱하여 품의서 양식의 **용도별분석** / **총괄경제** 표를 생성합니다.")

    if not working_files:
        st.info("👆 좌측 사이드바에 파일이 로드되지 않았습니다. 깃허브에 기초자료 파일을 올리거나 직접 업로드 해주세요.")
    else:
        # ── 기초자료 파싱 (탭1과 동일 로직 + 추가 컬럼) ──
        t3_clean_list = []
        for file_obj in working_files:
            try:
                file_name = file_obj.name if hasattr(file_obj, 'name') else os.path.basename(file_obj)
                if '현황' in file_name or '승인내역' in file_name or '품의서' in file_name:
                    continue
                if '투자' in file_name and '기초자료' not in file_name:
                    continue
                if hasattr(file_obj, 'seek'):
                    file_obj.seek(0)
                match = re.search(r'(\d+)차', file_name)
                cha_num = int(match.group(1)) if match else 1
                if file_name.endswith('.csv'):
                    df = pd.read_csv(file_obj, header=None, encoding='utf-8-sig')
                else:
                    df = pd.read_excel(file_obj, header=None)
                df = df.iloc[:100]
                idx_usage = 0
                idx_name = get_col_idx(df, ["구간명"], exact=True)
                idx_len = get_col_idx(df, ["길이(m)", "배관길이"], exact=False)
                idx_inv = get_col_idx(df, ["배관투자금액", "총공사비"], exact=False)
                idx_contrib = get_col_idx(df, ["시설분담금"], exact=False)
                idx_other = get_col_idx(df, ["기타이익", "보조금"], exact=False)
                idx_jeon = get_col_idx(df, ["수요전수계", "총전수"], exact=False)
                idx_jeon_apt = get_col_idx(df, ["공동주택전수"], exact=False)
                idx_jeon_single = get_col_idx(df, ["단독주택전수"], exact=False)
                idx_vol = get_col_idx(df, ["계(MJ)"], exact=False)
                idx_rev = get_col_idx(df, ["연간판매액", "판매액"], exact=False)
                idx_cost = get_col_idx(df, ["연간판매원가", "판매원가"], exact=False)
                idx_npv = get_col_idx(df, ["NPV"], exact=False)
                idx_irr = get_col_idx(df, ["IRR"], exact=False)
                # 추가 컬럼: 가스소비량, 취사전용, 수요가부담금
                idx_gas_vol = get_col_idx(df, ["가스소비량", "소비량합계"], exact=False)
                idx_dambu_cook = get_col_idx(df, ["취사전용"], exact=False)
                idx_dambu_demand = get_col_idx(df, ["수요가부담금"], exact=False)
                if idx_name is None:
                    continue
                md = {}
                md['차수'] = cha_num
                md['용도'] = df.iloc[:, idx_usage] if idx_usage is not None else '미분류'
                md['구간명'] = df.iloc[:, idx_name]
                md['길이'] = df.iloc[:, idx_len] if idx_len is not None else 0
                md['투자비'] = df.iloc[:, idx_inv] if idx_inv is not None else 0
                md['분담금'] = df.iloc[:, idx_contrib] if idx_contrib is not None else 0
                md['기타이익'] = df.iloc[:, idx_other] if idx_other is not None else 0
                md['총전수'] = df.iloc[:, idx_jeon] if idx_jeon is not None else 0
                md['공동주택전수'] = df.iloc[:, idx_jeon_apt] if idx_jeon_apt is not None else 0
                md['단독주택전수'] = df.iloc[:, idx_jeon_single] if idx_jeon_single is not None else 0
                md['판매량'] = df.iloc[:, idx_vol] if idx_vol is not None else 0
                md['판매액'] = df.iloc[:, idx_rev] if idx_rev is not None else 0
                md['판매원가'] = df.iloc[:, idx_cost] if idx_cost is not None else 0
                md['원본_NPV'] = df.iloc[:, idx_npv] if idx_npv is not None else 0
                md['원본_IRR'] = df.iloc[:, idx_irr] if idx_irr is not None else 0
                md['가스소비량'] = df.iloc[:, idx_gas_vol] if idx_gas_vol is not None else 0
                md['취사전용'] = df.iloc[:, idx_dambu_cook] if idx_dambu_cook is not None else 0
                md['수요가부담금'] = df.iloc[:, idx_dambu_demand] if idx_dambu_demand is not None else 0
                t3_clean_list.append(pd.DataFrame(md))
            except:
                pass

        if not t3_clean_list:
            st.warning("기초자료 파일에서 유효 데이터를 찾을 수 없습니다.")
        else:
            t3_df = pd.concat(t3_clean_list, ignore_index=True)
            t3_df['구간명'] = t3_df['구간명'].astype(str).str.strip()
            t3_df['용도'] = t3_df['용도'].astype(str).str.strip()
            t3_df['용도'] = t3_df['용도'].replace('ROE', '투자보수율가산')
            t3_df['용도'] = t3_df['용도'].replace('연료전지', '연료전지용')
            valid_usages_t3 = ["공공택지", "공동주택", "산업용", "업무용", "영업용", "연료전지용", "주택용", "주택용(지자체)", "투자보수율가산"]
            t3_df = t3_df[t3_df['용도'].isin(valid_usages_t3)]
            invalid_names = ['', '0', 'nan', 'none', 'null', '구간명', '소계', '합계', '총계', 'roe제외']
            t3_df = t3_df[~t3_df['구간명'].str.lower().isin(invalid_names)]
            t3_df = t3_df[~t3_df['구간명'].str.contains('합계|소계|총계|ROE', na=False, regex=True)]
            t3_df = t3_df.drop_duplicates(subset=['차수', '구간명'], keep='last')
            num_cols_t3 = ['길이', '투자비', '분담금', '기타이익', '총전수', '공동주택전수', '단독주택전수',
                           '판매량', '판매액', '판매원가', '원본_NPV', '원본_IRR',
                           '가스소비량', '취사전용', '수요가부담금']
            for c in num_cols_t3:
                if t3_df[c].dtype == object:
                    t3_df[c] = t3_df[c].astype(str).str.replace(',', '', regex=False)
                t3_df[c] = pd.to_numeric(t3_df[c], errors='coerce').fillna(0)
            t3_df['총전수'] = np.maximum(t3_df['총전수'], t3_df['공동주택전수'] + t3_df['단독주택전수'])
            t3_df['기본요금수익'] = 0.0
            tmp_u = t3_df['용도'].astype(str).str.replace(' ', '', regex=False)
            is_home_t3 = tmp_u.str.contains('주택|가정|공동|택지') & ~tmp_u.str.contains('외')
            t3_df.loc[is_home_t3, '기본요금수익'] = t3_df.loc[is_home_t3, '총전수'] * sim_basic_price_t3 * 12

            calc_cols_t3 = ['길이', '투자비', '분담금', '기타이익', '총전수', '판매량', '판매액', '판매원가',
                            '기본요금수익', '원본_NPV', '원본_IRR', '가스소비량', '취사전용', '수요가부담금']

            # ── 차수 선택 ──
            available_chas_t3 = sorted(t3_df['차수'].unique())
            selected_cha_t3 = st.selectbox("📌 기준 차수 선택", available_chas_t3, index=len(available_chas_t3)-1, format_func=lambda x: f"{x}차", key="t3_cha")
            filtered_t3 = t3_df[t3_df['차수'] == selected_cha_t3]
            st.info(f"**{selected_cha_t3}차** 당해차수 품의서를 생성합니다.")

            if filtered_t3.empty:
                st.warning(f"{selected_cha_t3}차 데이터가 없습니다.")
            else:
                # ── NPV/IRR/회수년수 계산 함수 (4값 반환) ──
                def t3_calc(row, usage_val=''):
                    npv_s, irr_s, irr_msg_s, flows_s = calculate_simulation(
                        row['길이'], row['투자비'], row['분담금'], row['기타이익'],
                        row['판매량'], row['판매액'], row['판매원가'],
                        row['총전수'], row['기본요금수익'],
                        RATE_T3, TAX_T3, dep_period_t3, analysis_period_t3,
                        c_maint_t3, c_adm_jeon_t3, c_adm_m_t3
                    )
                    if '공공택지' in str(usage_val) or '주택용' in str(usage_val) or '영업용' in str(usage_val):
                        npv = row['원본_NPV']
                        irr = (row['원본_IRR'] / 100.0) if row['원본_IRR'] != 0 else 0
                        irr_msg = ""
                    else:
                        npv = npv_s; irr = irr_s; irr_msg = irr_msg_s
                    # 회수년수 계산
                    payback = None
                    if flows_s:
                        cum = 0
                        for yi, fv in enumerate(flows_s):
                            cum += fv
                            if cum >= 0 and yi > 0:
                                payback = yi
                                break
                    return npv, irr, irr_msg, payback

                custom_order_t3 = ["공공택지", "공동주택", "산업용", "업무용", "영업용", "연료전지용", "주택용", "주택용(지자체)"]
                unique_usages_t3 = filtered_t3['용도'].unique().tolist()
                # 실적이 있는 용도만 (길이 or 투자비 > 0)
                usages_with_data = []
                for u in unique_usages_t3:
                    u_sub = filtered_t3[filtered_t3['용도'] == u]
                    if u_sub['길이'].sum() > 0 or u_sub['투자비'].sum() > 0:
                        usages_with_data.append(u)
                sorted_usages_t3 = sorted(usages_with_data, key=lambda x: 9999 if x == '투자보수율가산' else (custom_order_t3.index(x) if x in custom_order_t3 else 999))

                # ── 요약 카드 (차수 선택 바로 아래) ──
                _pre_ydb = []
                for u in sorted_usages_t3:
                    u_df_pre = filtered_t3[filtered_t3['용도'] == u]
                    u_sum_pre = u_df_pre[calc_cols_t3].sum()
                    u_npv_pre, u_irr_pre, _, _ = t3_calc(u_sum_pre, u)
                    _pre_ydb.append({'건수': len(u_df_pre), '배관투자금액': u_sum_pre['투자비'], 'NPV': u_npv_pre})
                _tot_cnt = sum(r['건수'] for r in _pre_ydb)
                _tot_inv = sum(r['배관투자금액'] for r in _pre_ydb)
                _tot_npv_pre = sum(r['NPV'] for r in _pre_ydb)
                _all_sum_pre = filtered_t3[filtered_t3['용도'].isin(usages_with_data)][calc_cols_t3].sum()
                _, _tot_irr_pre, _tot_irr_msg_pre, _ = t3_calc(_all_sum_pre, '합산')
                _irr_str = f"{_tot_irr_pre*100:.2f}%" if _tot_irr_pre is not None else _tot_irr_msg_pre

                st.markdown(
                    f"""
                    <div style="background: linear-gradient(135deg, #EBF5FB 0%, #D6EAF8 100%);
                                border-radius: 12px; padding: 20px 28px; margin-bottom: 24px;
                                border: 1px solid #AED6F1;">
                        <p style="margin:0 0 14px 0; font-size:17px; color:#2C3E50; font-weight:600;">
                            📋 {selected_cha_t3}차 품의서 요약
                        </p>
                        <div style="display:flex; justify-content:space-between; gap:16px; flex-wrap:wrap;">
                            <div style="flex:1; min-width:140px; text-align:center;">
                                <div style="font-size:16px; color:#5D6D7E;">총 건수</div>
                                <div style="font-size:32px; font-weight:700; color:#1A5276;">{_tot_cnt} <span style="font-size:18px; font-weight:400;">건</span></div>
                            </div>
                            <div style="flex:1; min-width:140px; text-align:center;">
                                <div style="font-size:16px; color:#5D6D7E;">총 배관투자금액</div>
                                <div style="font-size:32px; font-weight:700; color:#1A5276;">{_tot_inv:,.0f} <span style="font-size:18px; font-weight:400;">원</span></div>
                            </div>
                            <div style="flex:1; min-width:140px; text-align:center;">
                                <div style="font-size:16px; color:#5D6D7E;">합산 NPV</div>
                                <div style="font-size:32px; font-weight:700; color:#1A5276;">{_tot_npv_pre:,.0f} <span style="font-size:18px; font-weight:400;">원</span></div>
                            </div>
                            <div style="flex:1; min-width:140px; text-align:center;">
                                <div style="font-size:16px; color:#5D6D7E;">합산 IRR</div>
                                <div style="font-size:32px; font-weight:700; color:#1A5276;">{_irr_str}</div>
                            </div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True
                )
                st.markdown("---")

                # ══════════════════════════════════════════════════════
                # [섹션 1] 용도별분석
                # ══════════════════════════════════════════════════════
                st.subheader("1. 📊 신규 배관 투자 경제성 분석서 (용도별)")
                ydb_rows = []
                for u in sorted_usages_t3:
                    u_df = filtered_t3[filtered_t3['용도'] == u]
                    u_sum = u_df[calc_cols_t3].sum()
                    u_npv, u_irr, u_irr_msg, u_payback = t3_calc(u_sum, u)
                    건수 = len(u_df)
                    # IRR 표시: 100% 초과 시 "한도초과"
                    if u == '투자보수율가산':
                        irr_disp = '-'
                    elif u_irr is not None:
                        if u_irr * 100 > 100:
                            irr_disp = '한도초과'
                        else:
                            irr_disp = f"{u_irr*100:.2f}%"
                    else:
                        irr_disp = u_irr_msg if u_irr_msg else '-'
                    # 파생값 계산
                    판매수익 = u_sum['판매액'] - u_sum['판매원가']
                    환산세대 = (u_sum['총전수'] / u_sum['길이']) * 100 if u_sum['길이'] > 0 else 0
                    분담금일반 = u_sum['분담금'] - u_sum['취사전용'] - u_sum['수요가부담금']
                    if 분담금일반 < 0:
                        분담금일반 = u_sum['분담금']
                    ydb_rows.append({
                        '용도': u, '건수': 건수, '길이(m)': u_sum['길이'],
                        '배관투자금액(원)': u_sum['투자비'], '시설투자': 0, '총투자금액': u_sum['투자비'],
                        '전수(전)': u_sum['총전수'], '가스소비량': u_sum['가스소비량'],
                        '분담금일반': 분담금일반, '취사전용': u_sum['취사전용'],
                        '수요가부담금': u_sum['수요가부담금'], '총시설분담금': u_sum['분담금'],
                        '기타이익': u_sum['기타이익'],
                        '판매량(MJ/년)': u_sum['판매량'], '판매액(원/년)': u_sum['판매액'],
                        '판매원가(원/년)': u_sum['판매원가'], '판매수익(원/년)': 판매수익,
                        '회수년수': u_payback if u_payback is not None else '',
                        'NPV(원)': u_npv, 'IRR(%)': irr_disp,
                        '100m환산세대수': 환산세대,
                    })
                # 합계 행
                if ydb_rows:
                    all_filtered_sum = filtered_t3[filtered_t3['용도'].isin(usages_with_data)][calc_cols_t3].sum()
                    tot_npv_t3, tot_irr_t3, tot_irr_msg_t3, tot_payback_t3 = t3_calc(all_filtered_sum, '합산')
                    if tot_irr_t3 is not None:
                        if tot_irr_t3 * 100 > 100:
                            tot_irr_disp = '한도초과'
                        else:
                            tot_irr_disp = f"{tot_irr_t3*100:.2f}%"
                    else:
                        tot_irr_disp = tot_irr_msg_t3
                    tot_판매수익 = sum(r['판매수익(원/년)'] for r in ydb_rows)
                    tot_길이 = sum(r['길이(m)'] for r in ydb_rows)
                    tot_전수 = sum(r['전수(전)'] for r in ydb_rows)
                    tot_환산세대 = (tot_전수 / tot_길이) * 100 if tot_길이 > 0 else 0
                    ydb_rows.append({
                        '용도': '합계',
                        '건수': sum(r['건수'] for r in ydb_rows),
                        '길이(m)': tot_길이,
                        '배관투자금액(원)': sum(r['배관투자금액(원)'] for r in ydb_rows),
                        '시설투자': 0,
                        '총투자금액': sum(r['총투자금액'] for r in ydb_rows),
                        '전수(전)': tot_전수,
                        '가스소비량': sum(r['가스소비량'] for r in ydb_rows),
                        '분담금일반': sum(r['분담금일반'] for r in ydb_rows),
                        '취사전용': sum(r['취사전용'] for r in ydb_rows),
                        '수요가부담금': sum(r['수요가부담금'] for r in ydb_rows),
                        '총시설분담금': sum(r['총시설분담금'] for r in ydb_rows),
                        '기타이익': sum(r['기타이익'] for r in ydb_rows),
                        '판매량(MJ/년)': sum(r['판매량(MJ/년)'] for r in ydb_rows),
                        '판매액(원/년)': sum(r['판매액(원/년)'] for r in ydb_rows),
                        '판매원가(원/년)': sum(r['판매원가(원/년)'] for r in ydb_rows),
                        '판매수익(원/년)': tot_판매수익,
                        '회수년수': tot_payback_t3 if tot_payback_t3 is not None else '',
                        'NPV(원)': sum(r['NPV(원)'] for r in ydb_rows),
                        'IRR(%)': tot_irr_disp,
                        '100m환산세대수': tot_환산세대,
                    })
                    df_ydb = pd.DataFrame(ydb_rows)
                    # 화면 표시용 22컬럼 (엑셀과 동일)
                    disp_cols_ydb = ['용도', '건수', '길이(m)', '배관투자금액(원)', '시설투자', '총투자금액',
                                     '전수(전)', '가스소비량',
                                     '분담금일반', '취사전용', '수요가부담금', '총시설분담금',
                                     '기타이익',
                                     '판매량(MJ/년)', '판매액(원/년)', '판매원가(원/년)', '판매수익(원/년)',
                                     '회수년수', 'NPV(원)', 'IRR(%)', '100m환산세대수']
                    df_ydb_disp = df_ydb[disp_cols_ydb].copy()

                    def _sty_ydb(row):
                        if row['용도'] == '합계':
                            return ['background-color: #D0E8FF; font-weight: bold'] * len(row)
                        return [''] * len(row)

                    fmt_ydb = {c: '{:,.0f}' for c in disp_cols_ydb if c not in ['용도', 'IRR(%)', '회수년수']}
                    fmt_ydb['100m환산세대수'] = '{:,.2f}'
                    st.dataframe(df_ydb_disp.style.apply(_sty_ydb, axis=1).format(fmt_ydb),
                                 use_container_width=True, hide_index=True)

                st.markdown("<hr style='border-top: 2px solid #1e3a8a; margin: 40px 0 20px 0;'>", unsafe_allow_html=True)

                # ══════════════════════════════════════════════════════
                # [섹션 2] 총괄경제 (구간별)
                # ══════════════════════════════════════════════════════
                st.subheader("2. 📋 신규 배관 투자 경제성 분석서 (구간별 / 총괄경제)")

                tg_rows = []
                for u in sorted_usages_t3:
                    u_df = filtered_t3[filtered_t3['용도'] == u].sort_values('구간명')
                    for _, row in u_df.iterrows():
                        r_npv, r_irr, r_irr_msg, r_payback = t3_calc(row, u)
                        if u == '투자보수율가산':
                            irr_d = '-'
                        elif r_irr is not None:
                            if r_irr * 100 > 100:
                                irr_d = '한도초과'
                            else:
                                irr_d = f"{r_irr*100:.2f}%"
                        else:
                            irr_d = r_irr_msg if r_irr_msg else '-'
                        # 파생값
                        r_판매수익 = row['판매액'] - row['판매원가']
                        r_환산세대 = (row['총전수'] / row['길이']) * 100 if row['길이'] > 0 else 0
                        r_분담금일반 = row['분담금'] - row['취사전용'] - row['수요가부담금']
                        if r_분담금일반 < 0:
                            r_분담금일반 = row['분담금']
                        tg_rows.append({
                            '용도': u,
                            '공사명': row['구간명'],
                            '길이(m)': row['길이'],
                            '배관투자금액(원)': row['투자비'],
                            '시설투자': 0,
                            '총투자금액': row['투자비'],
                            '전수(전)': row['총전수'],
                            '가스소비량': row['가스소비량'],
                            '분담금일반': r_분담금일반,
                            '취사전용': row['취사전용'],
                            '수요가부담금': row['수요가부담금'],
                            '총시설분담금': row['분담금'],
                            '기타이익': row['기타이익'],
                            '판매량(MJ/년)': row['판매량'],
                            '판매액(원/년)': row['판매액'],
                            '판매원가(원/년)': row['판매원가'],
                            '판매수익(원/년)': r_판매수익,
                            '회수년수': r_payback if r_payback is not None else '',
                            'NPV(원)': r_npv,
                            'IRR(%)': irr_d,
                            '100m환산세대수': r_환산세대,
                        })

                if tg_rows:
                    df_tg = pd.DataFrame(tg_rows)
                    # 화면 표시용 22컬럼 (엑셀과 동일)
                    disp_cols_tg = ['용도', '공사명', '길이(m)', '배관투자금액(원)', '시설투자', '총투자금액',
                                    '전수(전)', '가스소비량',
                                    '분담금일반', '취사전용', '수요가부담금', '총시설분담금',
                                    '기타이익',
                                    '판매량(MJ/년)', '판매액(원/년)', '판매원가(원/년)', '판매수익(원/년)',
                                    '회수년수', 'NPV(원)', 'IRR(%)', '100m환산세대수']
                    df_tg_disp = df_tg[disp_cols_tg].copy()

                    # 용도 필터
                    avail_u = [u for u in sorted_usages_t3 if u in df_tg['용도'].values]
                    sel_u = st.multiselect("표시할 용도 필터 (전체 = 선택 없음)", avail_u, default=[], key="t3_filter")
                    df_tg_f = df_tg_disp[df_tg_disp['용도'].isin(sel_u)] if sel_u else df_tg_disp.copy()

                    section_colors = {
                        '공공택지': '#EBF5FB', '공동주택': '#EBF5FB', '산업용': '#FEF9E7',
                        '업무용': '#EAF7F0', '영업용': '#FDF2F8', '투자보수율가산': '#F2F3F4',
                    }
                    def _sty_tg(row):
                        c = section_colors.get(row['용도'], '')
                        return [f'background-color: {c}'] * len(row) if c else [''] * len(row)

                    fmt_tg = {c: '{:,.0f}' for c in disp_cols_tg if c not in ['용도', '공사명', 'IRR(%)', '회수년수']}
                    fmt_tg['100m환산세대수'] = '{:,.2f}'
                    st.dataframe(df_tg_f.style.apply(_sty_tg, axis=1).format(fmt_tg),
                                 use_container_width=True, hide_index=True)

                    # 용도별 소계 — df_ydb(용도별분석) 데이터를 그대로 사용 (이미 22컬럼 집계 완료)
                    st.markdown("#### 📌 용도별 소계")
                    disp_cols_sub = ['용도', '건수', '길이(m)', '배관투자금액(원)', '시설투자', '총투자금액',
                                     '전수(전)', '가스소비량',
                                     '분담금일반', '취사전용', '수요가부담금', '총시설분담금',
                                     '기타이익',
                                     '판매량(MJ/년)', '판매액(원/년)', '판매원가(원/년)', '판매수익(원/년)',
                                     '회수년수', 'NPV(원)', 'IRR(%)', '100m환산세대수']
                    # 필터 적용: 선택된 용도만 + 합계
                    if sel_u:
                        df_sub_disp = df_ydb[df_ydb['용도'].isin(sel_u + ['합계'])][disp_cols_sub].copy()
                    else:
                        df_sub_disp = df_ydb[disp_cols_sub].copy()

                    def _sty_agg(row):
                        if row['용도'] == '합계':
                            return ['background-color: #D6EAF8; font-weight: bold'] * len(row)
                        return [''] * len(row)

                    fmt_sub = {c: '{:,.0f}' for c in disp_cols_sub if c not in ['용도', 'IRR(%)', '회수년수']}
                    fmt_sub['100m환산세대수'] = '{:,.2f}'
                    st.dataframe(df_sub_disp.style.apply(_sty_agg, axis=1).format(fmt_sub),
                                 use_container_width=True, hide_index=True)

                # ══════════════════════════════════════════════════════
                # [섹션 3] B4 가로 엑셀 다운로드 (22컬럼 다단 헤더)
                # ══════════════════════════════════════════════════════
                st.markdown("<hr style='border-top: 2px solid #1e3a8a; margin: 40px 0 20px 0;'>", unsafe_allow_html=True)
                st.subheader("3. 📥 품의서 엑셀 다운로드 (B4 가로)")
                st.caption("B4 용지 가로 방향으로 출력 가능한 엑셀 파일을 다운로드합니다. 실적이 없는 용도는 자동 제외됩니다.")

                def generate_excel_b4(ydb_data, tg_data, cha_label):
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                    from openpyxl.utils import get_column_letter

                    wb = Workbook()
                    thin = Side(style='thin')
                    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
                    header_font = Font(name='맑은 고딕', bold=True, size=9)
                    title_font = Font(name='맑은 고딕', bold=True, size=14)
                    data_font = Font(name='맑은 고딕', size=9)
                    header_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
                    total_fill = PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid')
                    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    right_align = Alignment(horizontal='right', vertical='center')
                    left_align = Alignment(horizontal='left', vertical='center')
                    num_fmt = '#,##0'

                    # ── 22컬럼 정의 (용도별분석) ──
                    # col1=건수, col2=용도별, col3=길이, col4~6=투자금액(배관/시설/총),
                    # col7=수요전수계, col8=가스소비량, col9~12=시설분담금(일반/취사/수요가/총),
                    # col13=기타이익, col14~17=연간(판매량/판매액/판매원가/판매수익),
                    # col18~21=경제성분석(회수년수/NPV/IRR/100m환산), col22=비고
                    ydb_col_keys = [
                        '건수', '용도', '길이(m)',
                        '배관투자금액(원)', '시설투자', '총투자금액',
                        '전수(전)', '가스소비량',
                        '분담금일반', '취사전용', '수요가부담금', '총시설분담금',
                        '기타이익',
                        '판매량(MJ/년)', '판매액(원/년)', '판매원가(원/년)', '판매수익(원/년)',
                        '회수년수', 'NPV(원)', 'IRR(%)', '100m환산세대수',
                    ]

                    # ── 22컬럼 정의 (총괄경제) ──
                    # col1=용도별, col2=공사명, 나머지 col3~22 동일
                    tg_col_keys = [
                        '용도', '공사명', '길이(m)',
                        '배관투자금액(원)', '시설투자', '총투자금액',
                        '전수(전)', '가스소비량',
                        '분담금일반', '취사전용', '수요가부담금', '총시설분담금',
                        '기타이익',
                        '판매량(MJ/년)', '판매액(원/년)', '판매원가(원/년)', '판매수익(원/년)',
                        '회수년수', 'NPV(원)', 'IRR(%)', '100m환산세대수',
                    ]

                    def _setup_page(ws):
                        ws.page_setup.paperSize = 12  # B4 JIS
                        ws.page_setup.orientation = 'landscape'
                        ws.page_setup.fitToWidth = 1
                        ws.page_setup.fitToHeight = 0
                        ws.sheet_properties.pageSetUpPr.fitToPage = True
                        ws.page_margins.left = 0.3
                        ws.page_margins.right = 0.3
                        ws.page_margins.top = 0.4
                        ws.page_margins.bottom = 0.4

                    def _write_headers_ydb(ws):
                        """용도별분석 시트 - 22컬럼 다단 헤더 (row2=타이틀, row4=그룹, row5=세부, row6=단위)"""
                        ncols = 22
                        # Row 2: 타이틀
                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
                        c = ws.cell(row=2, column=1, value=f'신규 배관 투자 경제성 분석서(용도별) - {cha_label}')
                        c.font = title_font
                        c.alignment = center_align

                        # Row 4: 그룹 헤더 (merged)
                        # 독립 컬럼(row4~5 세로 병합): 건수(1), 용도별(2), 길이(3), 수요전수계(7), 가스소비량합계(8), 기타이익(13), 비고(22)
                        standalone_cols = {
                            1: '공급대상\n총계', 2: '용도별', 3: '길이',
                            7: '수요\n전수계', 8: '가스소비량\n합계', 13: '기타이익', 22: '비고'
                        }
                        for col_num, label in standalone_cols.items():
                            ws.merge_cells(start_row=4, start_column=col_num, end_row=5, end_column=col_num)
                            cell = ws.cell(row=4, column=col_num, value=label)
                            cell.font = header_font; cell.fill = header_fill
                            cell.border = border_all; cell.alignment = center_align

                        # 그룹 병합 헤더 (row4)
                        groups = [
                            (4, 6, '투자금액'),
                            (9, 12, '시설분담금'),
                            (14, 17, '연간'),
                            (18, 21, '경제성분석'),
                        ]
                        for start_c, end_c, label in groups:
                            ws.merge_cells(start_row=4, start_column=start_c, end_row=4, end_column=end_c)
                            cell = ws.cell(row=4, column=start_c, value=label)
                            cell.font = header_font; cell.fill = header_fill
                            cell.border = border_all; cell.alignment = center_align
                            # 나머지 셀 테두리
                            for ci in range(start_c+1, end_c+1):
                                ws.cell(row=4, column=ci).border = border_all
                                ws.cell(row=4, column=ci).fill = header_fill

                        # Row 5: 세부 헤더
                        detail_headers = {
                            4: '배관투자\n금액', 5: '시설투자\n금액', 6: '총투자\n금액',
                            9: '일반', 10: '취사전용', 11: '수요가\n부담금', 12: '총시설\n분담금',
                            14: '판매량', 15: '판매액', 16: '판매원가', 17: '판매수익',
                            18: '회수\n년수', 19: 'NPV', 20: 'IRR', 21: '100m\n환산세대수',
                        }
                        for col_num, label in detail_headers.items():
                            cell = ws.cell(row=5, column=col_num, value=label)
                            cell.font = header_font; cell.fill = header_fill
                            cell.border = border_all; cell.alignment = center_align

                        # Row 6: 단위 헤더
                        unit_headers = {
                            1: '', 2: '', 3: '(m)',
                            4: '(원)', 5: '(원)', 6: '(원)',
                            7: '(전)', 8: '(MJ/년)',
                            9: '(원)', 10: '(원)', 11: '(원)', 12: '(원)',
                            13: '(원)',
                            14: '(MJ/년)', 15: '(원/년)', 16: '(원/년)', 17: '(원/년)',
                            18: '(년)', 19: '(원)', 20: '(%)', 21: '(세대)',
                            22: '',
                        }
                        for col_num, label in unit_headers.items():
                            cell = ws.cell(row=6, column=col_num, value=label)
                            cell.font = Font(name='맑은 고딕', size=8)
                            cell.fill = header_fill
                            cell.border = border_all; cell.alignment = center_align

                        # 모든 헤더 행 테두리 보완
                        for r in [4, 5, 6]:
                            for ci in range(1, ncols+1):
                                ws.cell(row=r, column=ci).border = border_all

                    def _write_headers_tg(ws):
                        """총괄경제 시트 - 22컬럼 다단 헤더"""
                        ncols = 22
                        # Row 2: 타이틀
                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
                        c = ws.cell(row=2, column=1, value=f'신규 배관 투자 경제성 분석서(구간별) - {cha_label}')
                        c.font = title_font
                        c.alignment = center_align

                        # Row 4: 독립 컬럼 (세로 병합)
                        standalone_cols = {
                            1: '용도별', 2: '공사명', 3: '길이',
                            7: '수요\n전수계', 8: '가스소비량\n합계', 13: '기타이익', 22: '비고'
                        }
                        for col_num, label in standalone_cols.items():
                            ws.merge_cells(start_row=4, start_column=col_num, end_row=5, end_column=col_num)
                            cell = ws.cell(row=4, column=col_num, value=label)
                            cell.font = header_font; cell.fill = header_fill
                            cell.border = border_all; cell.alignment = center_align

                        # 그룹 병합 헤더
                        groups = [
                            (4, 6, '투자금액'),
                            (9, 12, '시설분담금'),
                            (14, 17, '연간'),
                            (18, 21, '경제성분석'),
                        ]
                        for start_c, end_c, label in groups:
                            ws.merge_cells(start_row=4, start_column=start_c, end_row=4, end_column=end_c)
                            cell = ws.cell(row=4, column=start_c, value=label)
                            cell.font = header_font; cell.fill = header_fill
                            cell.border = border_all; cell.alignment = center_align
                            for ci in range(start_c+1, end_c+1):
                                ws.cell(row=4, column=ci).border = border_all
                                ws.cell(row=4, column=ci).fill = header_fill

                        # Row 5: 세부 헤더
                        detail_headers = {
                            4: '배관투자\n금액', 5: '시설투자\n금액', 6: '총투자\n금액',
                            9: '일반', 10: '취사전용', 11: '수요가\n부담금', 12: '총시설\n분담금',
                            14: '판매량', 15: '판매액', 16: '판매원가', 17: '판매수익',
                            18: '회수\n년수', 19: 'NPV', 20: 'IRR', 21: '100m\n환산세대수',
                        }
                        for col_num, label in detail_headers.items():
                            cell = ws.cell(row=5, column=col_num, value=label)
                            cell.font = header_font; cell.fill = header_fill
                            cell.border = border_all; cell.alignment = center_align

                        # Row 6: 단위
                        unit_headers = {
                            1: '', 2: '', 3: '(m)',
                            4: '(원)', 5: '(원)', 6: '(원)',
                            7: '(전)', 8: '(MJ/년)',
                            9: '(원)', 10: '(원)', 11: '(원)', 12: '(원)',
                            13: '(원)',
                            14: '(MJ/년)', 15: '(원/년)', 16: '(원/년)', 17: '(원/년)',
                            18: '(년)', 19: '(원)', 20: '(%)', 21: '(세대)',
                            22: '',
                        }
                        for col_num, label in unit_headers.items():
                            cell = ws.cell(row=6, column=col_num, value=label)
                            cell.font = Font(name='맑은 고딕', size=8)
                            cell.fill = header_fill
                            cell.border = border_all; cell.alignment = center_align

                        for r in [4, 5, 6]:
                            for ci in range(1, ncols+1):
                                ws.cell(row=r, column=ci).border = border_all

                    def _write_data_row(ws, row_num, vals, is_total=False, is_subtotal=False):
                        """22개 값을 한 행에 씀"""
                        fill = None
                        if is_total:
                            fill = total_fill
                        elif is_subtotal:
                            fill = PatternFill(start_color='E8F8F5', end_color='E8F8F5', fill_type='solid')
                        for ci, v in enumerate(vals, 1):
                            cell = ws.cell(row=row_num, column=ci, value=v)
                            cell.font = Font(name='맑은 고딕', bold=(is_total or is_subtotal), size=9)
                            cell.border = border_all
                            if fill:
                                cell.fill = fill
                            # 정렬 및 포맷
                            if ci in (1, 2):  # 건수/용도 or 용도/공사명
                                cell.alignment = center_align
                            elif ci == 20:  # IRR
                                cell.alignment = right_align
                            elif ci == 18:  # 회수년수
                                cell.alignment = right_align
                                if isinstance(v, (int, float)):
                                    cell.number_format = '#,##0'
                            elif ci == 22:  # 비고
                                cell.alignment = left_align
                            else:
                                cell.alignment = right_align
                                if isinstance(v, (int, float)):
                                    cell.number_format = num_fmt

                    # ─── Sheet 1: 용도별분석 ───
                    ws1 = wb.active
                    ws1.title = '용도별분석'
                    _setup_page(ws1)
                    _write_headers_ydb(ws1)

                    # 데이터 (row 7부터)
                    data_start = 7
                    for ri, row_d in enumerate(ydb_data):
                        is_tot = row_d.get('용도', '') == '합계'
                        vals = [row_d.get(k, '') for k in ydb_col_keys]
                        vals.append('')  # 비고 (col22)
                        _write_data_row(ws1, data_start + ri, vals, is_total=is_tot)

                    # 열 너비
                    col_widths1 = [6, 12, 8, 14, 10, 14, 8, 12, 12, 10, 12, 12, 10, 12, 12, 12, 12, 6, 14, 8, 8, 8]
                    for ci, w in enumerate(col_widths1, 1):
                        ws1.column_dimensions[get_column_letter(ci)].width = w

                    # ─── Sheet 2: 총괄경제 ───
                    ws2 = wb.create_sheet('총괄경제')
                    _setup_page(ws2)
                    _write_headers_tg(ws2)

                    ncols = 22
                    section_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
                    subtotal_fill_xl = PatternFill(start_color='E8F8F5', end_color='E8F8F5', fill_type='solid')
                    current_row = 7

                    # 용도별로 그룹
                    grouped_tg = {}
                    for r in tg_data:
                        u = r['용도']
                        if u not in grouped_tg:
                            grouped_tg[u] = []
                        grouped_tg[u].append(r)

                    grand_sums = {k: 0 for k in tg_col_keys[2:] if k not in ('IRR(%)', '회수년수', '100m환산세대수')}
                    grand_count = 0

                    for u in sorted_usages_t3:
                        if u not in grouped_tg:
                            continue
                        rows_in_group = grouped_tg[u]

                        # 용도 섹션 헤더
                        ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ncols)
                        sec_cell = ws2.cell(row=current_row, column=1, value=f'■ {u} ({len(rows_in_group)}건)')
                        sec_cell.font = Font(name='맑은 고딕', bold=True, size=9)
                        sec_cell.fill = section_fill
                        sec_cell.border = border_all
                        sec_cell.alignment = left_align
                        for ci2 in range(2, ncols+1):
                            ws2.cell(row=current_row, column=ci2).border = border_all
                            ws2.cell(row=current_row, column=ci2).fill = section_fill
                        current_row += 1

                        # 소계 누적
                        sub_sums = {k: 0 for k in tg_col_keys[2:] if k not in ('IRR(%)', '회수년수', '100m환산세대수')}

                        for row_d in rows_in_group:
                            vals = [row_d.get(k, '') for k in tg_col_keys]
                            vals.append('')  # 비고(col22) - 일련번호 등
                            _write_data_row(ws2, current_row, vals)
                            # 소계 누적
                            for k in sub_sums:
                                try:
                                    sub_sums[k] += float(row_d.get(k, 0) or 0)
                                except:
                                    pass
                            current_row += 1

                        # 용도 소계 행
                        sub_전수 = sub_sums.get('전수(전)', 0)
                        sub_길이 = sub_sums.get('길이(m)', 0)
                        sub_환산 = (sub_전수 / sub_길이) * 100 if sub_길이 > 0 else 0
                        sub_vals = [
                            f'{u} 소계', '', sub_sums.get('길이(m)', 0),
                            sub_sums.get('배관투자금액(원)', 0), sub_sums.get('시설투자', 0), sub_sums.get('총투자금액', 0),
                            sub_sums.get('전수(전)', 0), sub_sums.get('가스소비량', 0),
                            sub_sums.get('분담금일반', 0), sub_sums.get('취사전용', 0),
                            sub_sums.get('수요가부담금', 0), sub_sums.get('총시설분담금', 0),
                            sub_sums.get('기타이익', 0),
                            sub_sums.get('판매량(MJ/년)', 0), sub_sums.get('판매액(원/년)', 0),
                            sub_sums.get('판매원가(원/년)', 0), sub_sums.get('판매수익(원/년)', 0),
                            '', '', '', sub_환산, '',
                        ]
                        _write_data_row(ws2, current_row, sub_vals, is_subtotal=True)
                        current_row += 1

                        grand_count += len(rows_in_group)
                        for k in grand_sums:
                            grand_sums[k] += sub_sums.get(k, 0)

                    # 총합계 행
                    gt_전수 = grand_sums.get('전수(전)', 0)
                    gt_길이 = grand_sums.get('길이(m)', 0)
                    gt_환산 = (gt_전수 / gt_길이) * 100 if gt_길이 > 0 else 0
                    gt_vals = [
                        '총합계', '', grand_sums.get('길이(m)', 0),
                        grand_sums.get('배관투자금액(원)', 0), grand_sums.get('시설투자', 0), grand_sums.get('총투자금액', 0),
                        grand_sums.get('전수(전)', 0), grand_sums.get('가스소비량', 0),
                        grand_sums.get('분담금일반', 0), grand_sums.get('취사전용', 0),
                        grand_sums.get('수요가부담금', 0), grand_sums.get('총시설분담금', 0),
                        grand_sums.get('기타이익', 0),
                        grand_sums.get('판매량(MJ/년)', 0), grand_sums.get('판매액(원/년)', 0),
                        grand_sums.get('판매원가(원/년)', 0), grand_sums.get('판매수익(원/년)', 0),
                        '', '', '', gt_환산, '',
                    ]
                    _write_data_row(ws2, current_row, gt_vals, is_total=True)

                    # 열 너비
                    col_widths2 = [12, 26, 8, 14, 10, 14, 8, 12, 12, 10, 12, 12, 10, 12, 12, 12, 12, 6, 14, 8, 8, 8]
                    for ci, w in enumerate(col_widths2, 1):
                        ws2.column_dimensions[get_column_letter(ci)].width = w

                    # 저장
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)
                    return output

                if ydb_rows and tg_rows:
                    cha_str = f"{selected_cha_t3}차"
                    excel_data = generate_excel_b4(ydb_rows, tg_rows, cha_str)
                    st.download_button(
                        label="📥 품의서 엑셀 다운로드 (B4 가로)",
                        data=excel_data,
                        file_name=f"품의서_경제성분석_{cha_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
